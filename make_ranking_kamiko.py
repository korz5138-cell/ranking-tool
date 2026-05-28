#!/usr/bin/env python3
"""差枚ランキング画像生成（神の子用：パステル・ポップデザイン）"""
import os
import re
import sys
import csv
import math
import glob
import unicodedata
from datetime import datetime
import openpyxl
from PIL import Image, ImageDraw, ImageFont, ImageFilter

_BUNDLED_VF = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), 'fonts', 'NotoSansJP-VF.ttf')


def _resolve_jp_fonts():
    """日本語フォントを weight 別に検出する。
    優先順位:
      1) macOS のヒラギノ（各 weight 別ファイル）
      2) リポジトリ同梱の Noto Sans JP variable font（"パス|weight名"形式）
      3) Linux: /usr/share/fonts 配下を glob で探索
    """
    mac = {
        'Black':  '/System/Library/Fonts/ヒラギノ角ゴシック W9.ttc',
        'Bold':   '/System/Library/Fonts/ヒラギノ角ゴシック W8.ttc',
        'Medium': '/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc',
        'Round':  '/System/Library/Fonts/ヒラギノ丸ゴ ProN W4.ttc',
    }
    if all(os.path.exists(p) for p in mac.values()):
        return mac

    # 同梱 variable font（"パス|weight名" で font() ヘルパが variation 適用）
    if os.path.exists(_BUNDLED_VF):
        return {
            'Black':  f'{_BUNDLED_VF}|Black',
            'Bold':   f'{_BUNDLED_VF}|Bold',
            'Medium': f'{_BUNDLED_VF}|Medium',
            'Round':  f'{_BUNDLED_VF}|Regular',
        }

    # OS の CJK フォントを glob で検索（最終手段）
    found = []
    for pat in (
        '/usr/share/fonts/opentype/noto/*CJK*.ttc',
        '/usr/share/fonts/truetype/noto/*CJK*.ttc',
        '/usr/share/fonts/opentype/noto-cjk/*.ttc',
        '/usr/share/fonts/**/NotoSans*JP*.otf',
        '/usr/share/fonts/**/NotoSansCJK*.otf',
        '/usr/share/fonts/**/*ipaex*.ttf',
    ):
        found.extend(glob.glob(pat, recursive=True))
    found = sorted(set(found))
    if found:
        def pick(*kws):
            for kw in kws:
                for p in found:
                    if kw.lower() in os.path.basename(p).lower():
                        return p
            return found[0]
        return {
            'Black':  pick('Black', 'Heavy', 'Bold'),
            'Bold':   pick('Bold', 'Black'),
            'Medium': pick('Medium', 'Regular'),
            'Round':  pick('Regular', 'Medium'),
        }
    return mac   # 最終的に PIL が cannot open resource を出す（デバッグしやすい）


_FONTS = _resolve_jp_fonts()
FONT_HEAVY = _FONTS['Black']
FONT_BOLD  = _FONTS['Bold']
FONT_MED   = _FONTS['Medium']
FONT_ROUND = _FONTS['Round']

def font(p, s):
    """ImageFont を生成。パスが '<path>|<variation>' 形式なら
    variable font の weight 軸を指定して読み込む。
    """
    p = str(p)
    if '|' in p:
        path, var_name = p.split('|', 1)
        f = ImageFont.truetype(path, s)
        try:
            f.set_variation_by_name(var_name)
        except Exception:
            pass   # 非 variable font の場合は無視して既定 weight
        return f
    return ImageFont.truetype(p, s)


# ===== 機種名マッピング =====
MODELS_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'models.csv')
_MODEL_ALIAS = None


def _normalize_model_key(s):
    """機種名照合用の正規化キー。
    NFKC（半角カナ→全角化、全角英数→半角化）後、
    区切り記号・装飾記号を除去 + 共通シリーズ接頭辞（パチスロ/スマスロ）除去 + 大小無視。
    これにより以下のような表記揺れが同一キーになる:
      'L東京ﾘﾍﾞﾝｼﾞｬｰｽﾞZF' ⇔ 'L 東京リベンジャーズ ZF'
      'LミリオンゴッドCX' ⇔ 'L/ミリオンゴッド/CX'
      'Lいざ!番長SB8' ⇔ 'L/いざ番長/SB8'
      'Lスーパーリオエース2ND02H' ⇔ 'Lパチスロスーパーリオエース2ND02H'
    """
    if s is None:
        return ''
    n = unicodedata.normalize('NFKC', str(s))
    for ch in (' ', '　', '/', '\\', '.', '-', '・', '~', '‾', '!', '?', '_',
               '(', ')', '（', '）', '「', '」', '『', '』', '【', '】', ',', '、'):
        n = n.replace(ch, '')
    # 先頭の L/S/LB プレフィックス直後にある「パチスロ/スマスロ/パチンコ」を除去
    for prefix in ('パチスロ', 'スマスロ', 'パチンコ'):
        idx = n.find(prefix)
        if 0 <= idx <= 2:
            n = n[:idx] + n[idx + len(prefix):]
    return n.lower().strip()


def _load_model_alias():
    """models.csv を読み、型式名 → 表示名 のマップを返す。
    優先順: 短縮名 > 機種名（短縮名が空の場合）。
    キーは _normalize_model_key で正規化。
    """
    global _MODEL_ALIAS
    if _MODEL_ALIAS is not None:
        return _MODEL_ALIAS
    table = {}
    if os.path.exists(MODELS_CSV):
        with open(MODELS_CSV, encoding='utf-8-sig', newline='') as f:
            reader = csv.DictReader(f)
            for r in reader:
                official = (r.get('機種名') or '').strip()
                code = (r.get('型式名') or '').strip()
                short = (r.get('短縮名') or '').strip()
                if not code:
                    continue
                disp = short or official
                if not disp:
                    continue
                key = _normalize_model_key(code)
                if key and key not in table:
                    table[key] = disp
                # 型式名以外に「機種名」自体でも引けるようにしておく（保険）
                key2 = _normalize_model_key(official)
                if key2 and key2 not in table:
                    table[key2] = disp
    _MODEL_ALIAS = table
    return table


def resolve_model_name(raw):
    """型式名（筐体名）を models.csv の短縮名/機種名に解決。
    1. 正規化キーで完全一致
    2. 「prefix(L/S/LB等の最初の2文字) + suffix(末尾の機種コード2〜4文字)」での一意絞り込み
    3. それでも見つからなければ簡易ヒューリスティック clean_studio_name
    """
    if raw is None:
        return ''
    table = _load_model_alias()
    key = _normalize_model_key(raw)
    if key in table:
        return table[key]
    # フォールバック: 先頭2文字 + 末尾2〜4文字一致での候補絞り込み
    # 例: Lかぐや様jA (lかぐや様ja) → Lかぐや様は告らせたいjA (lかぐや様は告らせたいja)
    if len(key) >= 5:
        head = key[:2]
        for suf_len in (4, 3, 2):
            if len(key) <= suf_len + 2:
                continue
            tail = key[-suf_len:]
            cands = set()
            for k, v in table.items():
                if k.startswith(head) and k.endswith(tail):
                    cands.add(v)
            if len(cands) == 1:
                return next(iter(cands))
    return clean_studio_name(unicodedata.normalize('NFKC', str(raw)).strip())


# ===== データ読込 =====
def short_name(n):
    s = str(n)
    for p in ('スマスロ ', 'スマスロ', 'Lパチスロ ', 'L パチスロ ', 'パチスロ ', 'L ', '真打 '):
        if s.startswith(p):
            return s[len(p):]
    return s


def load_ranking_xlsx(xlsx):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
            if r and r[0] is not None]
    rows.sort(key=lambda r: -(r[3] or 0))
    return [{
        'dai':   int(r[0]),
        'name':  short_name(r[2] if r[2] is not None else r[1]),
        'samai': int(r[3] or 0),
        'bb':    int(r[4] or 0),
        'rb':    int(r[5] or 0),
    } for r in rows[:10]]


def load_ranking_csv(csv_path):
    rows = []
    # 一部のCSVは先頭ヘッダー行のみバイナリ化け（ダウンロード時の破損等）。
    # その場合は標準ヘッダーで読み直して救済する。
    DEFAULT_HEADER = '機種名,台番,総ゲーム,BB,RB,ART,合成確率,差枚数'
    text = None
    try:
        with open(csv_path, encoding='utf-8-sig') as f:
            text = f.read()
    except UnicodeDecodeError:
        with open(csv_path, encoding='utf-8', errors='replace') as f:
            text = f.read()
    first, _, rest = text.partition('\n')
    expected_cols = ('機種名', '台番', '差枚')
    if not all(c in first for c in expected_cols):
        text = DEFAULT_HEADER + '\n' + rest
    import io
    reader = csv.DictReader(io.StringIO(text))
    for r in reader:
        try:
            dai = int(r.get('台番') or r.get('台番号') or 0)
            sa = int(r.get('差枚数') or r.get('差枚') or 0)
        except ValueError:
            continue
        if dai == 0:
            continue
        def to_int(x):
            try: return int(x)
            except (TypeError, ValueError): return 0
        rows.append({
            'dai':   dai,
            'name':  short_name(r.get('機種名') or ''),
            'samai': sa,
            'bb':    to_int(r.get('BB')),
            'rb':    to_int(r.get('RB')),
        })
    rows.sort(key=lambda x: -x['samai'])
    return rows[:10]


def clean_studio_name(n):
    s = unicodedata.normalize('NFKC', str(n))
    # 末尾のメーカーコード（区切り付き）を除去
    s = re.sub(r'[\s/][A-Z][A-Z0-9]{1,4}$', '', s)
    # 先頭の機種カテゴリプレフィックス
    s = re.sub(r'^L[B]?[/\s]+', '', s)
    s = re.sub(r'^S[/\s]+', '', s)
    for p in ('スマスロ', 'パチスロ', '真打'):
        if s.startswith(p):
            s = s[len(p):]
    return s.strip(' /')


def _classify_diff_label(label: str) -> str | None:
    """差分列ラベルが「プレイヤー視点（セーフ-アウト, +=客勝ち）」か
    「店側視点（アウト-セーフ, +=店勝ち）」かを判定。
    括弧書きの注釈（"(自動計算)" 等）は除去して判定する。
    return: 'player' / 'shop' / None
    """
    if not label:
        return None
    s = re.sub(r'[（(].*?[)）]', '', str(label)).strip()
    if s in ('差枚', '差枚数'):
        return 'player'   # セーフ - アウト
    if s in ('差', '差１', '差1'):
        return 'shop'     # アウト - セーフ
    return None


# シート走査の安全上限。Excel の空シートは最大 1,048,576 行を主張するため、
# 上限を設けないと巨大シートで著しいメモリ・時間消費が発生してクラッシュ要因になる。
_MAX_HEADER_SCAN = 200    # ヘッダー検索: 通常 1-20 行以内に存在
_MAX_DATA_SCAN = 5000     # データ行: 実際のホール台数の妥当な上限


def _find_header_columns(ws):
    """シート内から '台番' '機種名' '差(枚)' を含むヘッダー行を検出し、
    (header_row_1based, col_dai, col_name, col_diff, col_bb, col_rb, diff_kind) を返す。
    diff_kind は 'player' / 'shop'。検出失敗時は (None,)*7。
    """
    BB_LABELS = {'BB', 'BB回数', 'ＢＢ'}
    RB_LABELS = {'RB', 'RB回数', 'ＲＢ'}
    for i, r in enumerate(ws.iter_rows(max_row=_MAX_HEADER_SCAN, values_only=True), start=1):
        if not r:
            continue
        cells = list(r)
        col_dai = col_name = col_diff = col_bb = col_rb = None
        diff_kind = None
        for j, c in enumerate(cells):
            if not isinstance(c, str):
                continue
            cs = c.strip()
            if cs in ('台番', '台番号') and col_dai is None:
                col_dai = j
            elif cs == '機種名' and col_name is None:
                col_name = j
            elif cs in BB_LABELS and col_bb is None:
                col_bb = j
            elif cs in RB_LABELS and col_rb is None:
                col_rb = j
            elif col_diff is None:
                kind = _classify_diff_label(cs)
                if kind is not None:
                    col_diff = j
                    diff_kind = kind
        if col_dai is not None and col_name is not None and col_diff is not None:
            return i, col_dai, col_name, col_diff, col_bb, col_rb, diff_kind
    return None, None, None, None, None, None, None


def _count_data_rows(ws, header_row, col_dai, col_name):
    """ヘッダーより下にある有効データ行の数をカウント。
    Excel の空シートは max_row が極端に大きい値（〜100万行）を返すため、
    安全上限 _MAX_DATA_SCAN を超える行はスキャンしない。
    """
    max_col = max(col_dai, col_name)
    count = 0
    end_row = header_row + _MAX_DATA_SCAN
    for r in ws.iter_rows(min_row=header_row + 1, max_row=end_row, values_only=True):
        if not r or len(r) <= max_col:
            continue
        if r[col_dai] is None or r[col_name] is None:
            continue
        try:
            int(r[col_dai])
            count += 1
        except (TypeError, ValueError):
            pass
    return count


def _to_int(v):
    if v is None:
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return 0


def load_ranking_studio_xlsx(path):
    """神の子取材系 xlsx ローダー（複数シート・列位置のばらつきに対応）

    シート選択:
      - 'S' シートが存在し有効データを持つなら最優先
      - そうでなければ「有効データ行が最多」のシートを選択
    差枚の符号:
      - 列ラベルが '差' / '差１' / '差1' → 店側視点（アウト-セーフ）→ プレイヤー視点に反転
      - 列ラベルが '差枚' / '差枚数' / '差枚(自動計算)' 等 → 既にプレイヤー視点 → そのまま
    BB/RB 列があれば抽出。
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # 各シートを評価
    candidates = []  # (priority, count, ws, header_row, cd, cn, cf, cb, cr, diff_kind)
    for sn in wb.sheetnames:
        ws = wb[sn]
        hr, cd, cn, cf, cb, cr, dk = _find_header_columns(ws)
        if hr is None:
            continue
        count = _count_data_rows(ws, hr, cd, cn)
        priority = 2 if sn == 'S' else 1
        candidates.append((priority, count, ws, hr, cd, cn, cf, cb, cr, dk))

    if not candidates:
        raise ValueError(
            f'ヘッダー行を検出できません（台番・機種名・差/差枚 の3列が見つからない）: '
            f'{os.path.basename(path)}')

    # priority 降順 → count 降順
    candidates.sort(key=lambda x: (-x[0], -x[1]))
    (_, _, chosen_ws, header_row, col_dai, col_name, col_diff,
     col_bb, col_rb, diff_kind) = candidates[0]

    max_col = max(c for c in (col_dai, col_name, col_diff, col_bb, col_rb) if c is not None)
    rows = []
    end_row = header_row + _MAX_DATA_SCAN
    for r in chosen_ws.iter_rows(min_row=header_row + 1, max_row=end_row, values_only=True):
        if not r or len(r) <= max_col:
            continue
        dai_v = r[col_dai]
        name_v = r[col_name]
        diff_v = r[col_diff]
        if dai_v is None or name_v is None:
            continue
        if isinstance(name_v, str) and name_v.strip() in ('停止台', ''):
            continue
        try:
            dai = int(dai_v)
        except (TypeError, ValueError):
            continue
        raw_diff = _to_int(diff_v)
        # 列ラベルの符号規約に合わせてプレイヤー視点に統一
        samai = raw_diff if diff_kind == 'player' else -raw_diff
        row = {
            'dai':   dai,
            'name':  resolve_model_name(name_v),
            'samai': samai,
        }
        if col_bb is not None:
            row['bb'] = _to_int(r[col_bb])
        if col_rb is not None:
            row['rb'] = _to_int(r[col_rb])
        rows.append(row)
    rows.sort(key=lambda x: -x['samai'])
    return rows[:10]


def load_ranking(path):
    base = os.path.basename(path)
    # 神の子取材系xlsx（【店舗】開始 or ファイル名に「神の子」「取材」「調査」を含む）
    if base.lower().endswith('.xlsx') and (
        base.startswith('【') or '神の子' in base or '取材' in base or '調査' in base
    ):
        return load_ranking_studio_xlsx(path)
    if path.lower().endswith('.csv'):
        return load_ranking_csv(path)
    # 標準xlsx形式 → 失敗時は神の子形式にフォールバック
    try:
        result = load_ranking_xlsx(path)
        if result:
            return result
    except Exception:
        pass
    return load_ranking_studio_xlsx(path)


# ===== 装飾パーツ =====
def draw_star(draw, cx, cy, r, fill, outline=None, points=5):
    pts = []
    for i in range(points * 2):
        ang = math.radians(-90 + 180 * i / points)
        rad = r if i % 2 == 0 else r * 0.45
        pts.append((cx + rad * math.cos(ang), cy + rad * math.sin(ang)))
    draw.polygon(pts, fill=fill, outline=outline)


def draw_heart(draw, cx, cy, size, fill):
    """簡易ハート"""
    r = size // 2
    draw.ellipse([cx - size, cy - size, cx, cy], fill=fill)
    draw.ellipse([cx, cy - size, cx + size, cy], fill=fill)
    pts = [(cx - size + 2, cy - r // 2), (cx + size - 2, cy - r // 2),
           (cx, cy + size + r // 2)]
    draw.polygon(pts, fill=fill)


def draw_cloud(img, cx, cy, w, h, fill):
    """ふんわり雲（楕円の合成）"""
    layer = Image.new('RGBA', img.size, (0, 0, 0, 0))
    d = ImageDraw.Draw(layer)
    d.ellipse([cx - w / 2, cy - h / 2, cx + w / 2, cy + h / 2], fill=fill)
    d.ellipse([cx - w / 2 - w * 0.25, cy - h / 4, cx, cy + h / 2], fill=fill)
    d.ellipse([cx, cy - h / 4, cx + w / 2 + w * 0.25, cy + h / 2], fill=fill)
    d.ellipse([cx - w / 4, cy - h / 2 - h * 0.3, cx + w / 4, cy + h / 4], fill=fill)
    img.alpha_composite(layer)


# ===== 画像生成 =====
def render_image(ranking, date_full, store, out_path):
    yyyy = date_full[:4]
    mm = int(date_full[4:6])
    dd = int(date_full[6:8])
    date_badge = f'{mm}/{dd}'
    footer_date = f'{yyyy}年{mm}月{dd}日'

    # ---------- レイアウト ----------
    PAD = 12
    has_bb_rb = bool(ranking) and ('bb' in ranking[0])
    COLS = [
        ('rank', '順位',   100),
        ('dai',  '台番号', 170),
        ('name', '機種名', 320 if has_bb_rb else 470),
        ('sa',   '差枚',   200),
    ]
    if has_bb_rb:
        COLS.append(('bb', 'BB', 115))
        COLS.append(('rb', 'RB', 115))
    W = sum(c[2] for c in COLS) + PAD * 2
    HEADER_H = 220
    ROW_H = 70
    N = len(ranking)
    FOOTER_H = 90
    H = HEADER_H + ROW_H * N + FOOTER_H + 16

    # ---------- 背景：パステルピンク→水色のグラデ ----------
    img = Image.new('RGBA', (W, H), (255, 255, 255, 255))
    bg = Image.new('RGBA', (W, H))
    pix = bg.load()
    for y in range(H):
        t = y / max(H - 1, 1)
        # ピンク (255,228,236) → ラベンダー (235,225,255) → 水色 (220,240,255)
        if t < 0.5:
            tt = t / 0.5
            r = int(255 + (235 - 255) * tt)
            g = int(228 + (225 - 228) * tt)
            b = int(236 + (255 - 236) * tt)
        else:
            tt = (t - 0.5) / 0.5
            r = int(235 + (220 - 235) * tt)
            g = int(225 + (240 - 225) * tt)
            b = int(255 + (255 - 255) * tt)
        for x in range(W):
            pix[x, y] = (r, g, b, 255)
    img.alpha_composite(bg)

    # 装飾：ふわふわ雲（背景に薄く）
    draw_cloud(img, 90, 60, 120, 50, (255, 255, 255, 160))
    draw_cloud(img, W - 90, 130, 130, 55, (255, 255, 255, 130))
    draw_cloud(img, W - 150, H - 160, 110, 45, (255, 255, 255, 140))
    draw_cloud(img, 80, H - 100, 100, 40, (255, 255, 255, 150))

    draw = ImageDraw.Draw(img)

    # 装飾：星・ハート（ヘッダー周辺）
    draw_star(draw, 60, 50, 14, (255, 215, 100, 255), outline=(255, 180, 40))
    draw_star(draw, W - 50, 60, 10, (255, 215, 100, 255), outline=(255, 180, 40))
    draw_star(draw, W - 30, 180, 8, (255, 215, 100, 255), outline=(255, 180, 40))
    draw_heart(draw, 30, 170, 10, (255, 150, 180, 255))
    draw_heart(draw, W - 100, 30, 8, (255, 150, 180, 255))

    # ---------- ヘッダー：タイトル ----------
    # タイトル背景：白の角丸＋ピンクの縁
    title_box = (60, 30, W - 230, 150)
    # 影
    shadow = Image.new('RGBA', img.size, (0, 0, 0, 0))
    sd = ImageDraw.Draw(shadow)
    sd.rounded_rectangle((title_box[0] + 6, title_box[1] + 8,
                          title_box[2] + 6, title_box[3] + 8),
                         radius=30, fill=(255, 150, 180, 60))
    shadow = shadow.filter(ImageFilter.GaussianBlur(4))
    img.alpha_composite(shadow)
    draw.rounded_rectangle(title_box, radius=30,
                           fill=(255, 255, 255, 235),
                           outline=(255, 150, 180), width=4)

    # タイトル文字（ピンク+水色のツートン感を出すため2段）
    ft_sub  = font(FONT_HEAVY, 26)
    ft_main = font(FONT_HEAVY, 60)
    title_cx = (title_box[0] + title_box[2]) / 2
    draw.text((title_cx, 60), '☆ Slot ☆', font=ft_sub,
              fill=(120, 180, 230), anchor='mm',
              stroke_width=2, stroke_fill=(255, 255, 255))
    # メインタイトル：黒の縁取り＋ピンク
    draw.text((title_cx, 110), '差枚ランキング', font=ft_main,
              fill=(255, 105, 160), anchor='mm',
              stroke_width=4, stroke_fill=(255, 255, 255))

    # 日付バッジ（右上）：吹き出し風円
    bx, by, br = W - 130, 90, 70
    # 影
    shadow2 = Image.new('RGBA', img.size, (0, 0, 0, 0))
    sd2 = ImageDraw.Draw(shadow2)
    sd2.ellipse([bx - br + 4, by - br + 6, bx + br + 4, by + br + 6],
                fill=(120, 180, 230, 70))
    shadow2 = shadow2.filter(ImageFilter.GaussianBlur(4))
    img.alpha_composite(shadow2)
    draw.ellipse([bx - br, by - br, bx + br, by + br],
                 fill=(255, 255, 255, 245),
                 outline=(120, 180, 230), width=4)
    # 内側の点線風飾り
    draw.ellipse([bx - br + 8, by - br + 8, bx + br - 8, by + br - 8],
                 outline=(255, 200, 220), width=2)
    draw.text((bx, by - 18), '集計日', font=font(FONT_BOLD, 16),
              fill=(120, 180, 230), anchor='mm')
    draw.text((bx, by + 16), date_badge, font=font(FONT_HEAVY, 36),
              fill=(255, 105, 160), anchor='mm')

    # ---------- テーブルヘッダー ----------
    hb_y = HEADER_H - 50
    # ヘッダー帯：水色
    draw.rounded_rectangle([PAD, hb_y, W - PAD, hb_y + 44], radius=14,
                           fill=(160, 215, 245, 255),
                           outline=(120, 180, 230), width=3)
    f_head = font(FONT_HEAVY, 22)
    cx = PAD
    for key, lbl, w in COLS:
        draw.text((cx + w / 2, hb_y + 22), lbl, font=f_head,
                  fill=(60, 90, 130), anchor='mm',
                  stroke_width=2, stroke_fill=(255, 255, 255))
        cx += w
        if cx < W - PAD:
            # 区切り：白い丸ドット
            for dy in range(hb_y + 8, hb_y + 38, 8):
                draw.ellipse([cx - 1.5, dy, cx + 1.5, dy + 3],
                             fill=(255, 255, 255, 200))

    # ---------- ランクスタイル ----------
    def rank_style(rk):
        # bg/border/text-on-badge
        if rk == 1: return (255, 200, 80),  (255, 150, 30),  (30, 50, 110)    # ゴールド + 紺文字
        if rk == 2: return (220, 220, 235), (170, 170, 200), (90, 90, 130)    # シルバー
        if rk == 3: return (240, 190, 150), (200, 130, 80),  (110, 60, 30)    # ブロンズ
        return (255, 175, 200), (240, 130, 170), (160, 200, 240)              # ピンク + 薄青文字

    def row_bg(rk, even):
        if rk == 1: return (255, 245, 200, 230)   # ゴールド薄
        if rk == 2: return (240, 240, 255, 230)   # シルバー薄
        if rk == 3: return (255, 235, 215, 230)   # ブロンズ薄
        return (255, 255, 255, 230) if even else (250, 240, 250, 230)

    # ---------- 行 ----------
    f_rk     = font(FONT_HEAVY, 36)
    f_rk_lbl = font(FONT_BOLD, 17)
    f_dai    = font(FONT_HEAVY, 34)
    f_dai_lb = font(FONT_BOLD, 14)
    f_nm_l   = [font(FONT_HEAVY, s) for s in (28, 24, 20, 17)]
    f_val    = font(FONT_HEAVY, 32)
    f_unit   = font(FONT_BOLD, 14)
    f_count  = font(FONT_HEAVY, 30)

    def split_two(text, fnt, maxw):
        sep = [i for i, c in enumerate(text) if c in ' 　・-‐ｰー']
        best = None
        for i in sep:
            left = text[:i].rstrip()
            right = text[i + 1:].lstrip()
            if (draw.textlength(left, font=fnt) <= maxw and
                    draw.textlength(right, font=fnt) <= maxw):
                best = (left, right)
        if best:
            return best
        mid = len(text) // 2
        for off in range(0, len(text)):
            for k in (mid - off, mid + off):
                if 0 < k < len(text):
                    L, R = text[:k], text[k:]
                    if (draw.textlength(L, font=fnt) <= maxw and
                            draw.textlength(R, font=fnt) <= maxw):
                        return L, R
        right = text[mid:]
        while draw.textlength(right + '…', font=fnt) > maxw and len(right) > 1:
            right = right[:-1]
        return text[:mid], right + '…'

    y = HEADER_H
    for i, item in enumerate(ranking):
        rk = i + 1
        # 行背景：角丸の白パネル
        panel = Image.new('RGBA', img.size, (0, 0, 0, 0))
        pd = ImageDraw.Draw(panel)
        bg_c = row_bg(rk, i % 2 == 0)
        pd.rounded_rectangle([PAD, y + 2, W - PAD, y + ROW_H - 4],
                             radius=14, fill=bg_c,
                             outline=(255, 200, 220, 180) if rk > 3 else None,
                             width=2)
        img.alpha_composite(panel)

        cx = PAD
        # 順位バッジ（円）
        w = 100
        bgc, oc, tc = rank_style(rk)
        bcx = cx + w / 2
        bcy = y + ROW_H / 2
        br_ = 26
        # 影
        sh = Image.new('RGBA', img.size, (0, 0, 0, 0))
        sd_ = ImageDraw.Draw(sh)
        sd_.ellipse([bcx - br_ + 2, bcy - br_ + 4,
                     bcx + br_ + 2, bcy + br_ + 4],
                    fill=(180, 100, 140, 80))
        sh = sh.filter(ImageFilter.GaussianBlur(2))
        img.alpha_composite(sh)
        draw.ellipse([bcx - br_, bcy - br_, bcx + br_, bcy + br_],
                     fill=bgc, outline=oc, width=3)
        # ランク数字
        rs = str(rk)
        rs_size = 30 if rk < 10 else 24
        f_rk_ = font(FONT_HEAVY, rs_size)
        rsw = draw.textlength(rs, font=f_rk_)
        draw.text((bcx - 5, bcy - 1), rs, font=f_rk_, fill=tc, anchor='mm',
                  stroke_width=2, stroke_fill=(255, 255, 255))
        # 「位」
        draw.text((bcx + rsw / 2 + 3, bcy + 6), '位', font=f_rk_lbl,
                  fill=tc, anchor='lm',
                  stroke_width=2, stroke_fill=(255, 255, 255))
        # トップ3にだけ星マーク
        if rk <= 3:
            star_color = {(1): (255, 215, 60),
                          (2): (200, 200, 220),
                          (3): (220, 150, 90)}[rk]
            draw_star(draw, bcx + br_ - 4, bcy - br_ + 4, 9,
                      star_color, outline=(255, 255, 255))
        cx += w

        # 台番号
        w = 170
        s = str(item['dai'])
        sw = draw.textlength(s, font=f_dai)
        nx = cx + w / 2 - 14
        draw.text((nx, y + ROW_H / 2 - 1), s, font=f_dai,
                  fill=(90, 110, 160), anchor='mm',
                  stroke_width=2, stroke_fill=(255, 255, 255))
        draw.text((nx + sw / 2 + 4, y + ROW_H / 2 + 10), '番台',
                  font=f_dai_lb, fill=(255, 130, 170), anchor='lm')
        cx += w

        # 機種名
        w = 320 if has_bb_rb else 470
        nm = item['name']
        avail = w - 18
        one_line = None
        for fnt in f_nm_l:
            if draw.textlength(nm, font=fnt) <= avail:
                one_line = fnt
                break
        if one_line is not None:
            draw.text((cx + 12, y + ROW_H / 2), nm, font=one_line,
                      fill=(70, 70, 100), anchor='lm',
                      stroke_width=2, stroke_fill=(255, 255, 255))
        else:
            two_fnt = f_nm_l[-1]
            l1, l2 = split_two(nm, two_fnt, avail)
            draw.text((cx + 12, y + ROW_H / 2 - 14), l1, font=two_fnt,
                      fill=(70, 70, 100), anchor='lm',
                      stroke_width=2, stroke_fill=(255, 255, 255))
            draw.text((cx + 12, y + ROW_H / 2 + 14), l2, font=two_fnt,
                      fill=(70, 70, 100), anchor='lm',
                      stroke_width=2, stroke_fill=(255, 255, 255))
        cx += w

        # 差枚
        w = 200
        sa = item['samai']
        sa_color = (255, 90, 130) if sa >= 0 else (130, 160, 200)
        sa_text = ('+' if sa > 0 else '') + f'{sa:,}'
        mai_x = cx + w - 14
        draw.text((mai_x, y + ROW_H / 2 + 10), '枚', font=f_unit,
                  fill=(255, 130, 170), anchor='rm')
        draw.text((mai_x - 22, y + ROW_H / 2), sa_text, font=f_val,
                  fill=sa_color, anchor='rm',
                  stroke_width=2, stroke_fill=(255, 255, 255))
        cx += w

        # BB
        if not has_bb_rb:
            y += ROW_H
            continue
        w = 115
        bb_s = str(item['bb'])
        bbw = draw.textlength(bb_s, font=f_count)
        bb_cx = cx + w / 2 - 8
        draw.text((bb_cx, y + ROW_H / 2), bb_s, font=f_count,
                  fill=(120, 180, 230), anchor='mm',
                  stroke_width=2, stroke_fill=(255, 255, 255))
        draw.text((bb_cx + bbw / 2 + 3, y + ROW_H / 2 + 10), '回',
                  font=f_unit, fill=(255, 130, 170), anchor='lm')
        cx += w

        # RB
        w = 115
        rb_s = str(item['rb'])
        rbw = draw.textlength(rb_s, font=f_count)
        rb_cx = cx + w / 2 - 8
        draw.text((rb_cx, y + ROW_H / 2), rb_s, font=f_count,
                  fill=(120, 180, 230), anchor='mm',
                  stroke_width=2, stroke_fill=(255, 255, 255))
        draw.text((rb_cx + rbw / 2 + 3, y + ROW_H / 2 + 10), '回',
                  font=f_unit, fill=(255, 130, 170), anchor='lm')
        cx += w

        y += ROW_H

    # ---------- フッター ----------
    fy = y + 8
    # 角丸ピンクパネル
    panel = Image.new('RGBA', img.size, (0, 0, 0, 0))
    pd = ImageDraw.Draw(panel)
    pd.rounded_rectangle([PAD, fy, W - PAD, fy + FOOTER_H - 4], radius=20,
                         fill=(255, 220, 235, 240),
                         outline=(255, 150, 180), width=4)
    img.alpha_composite(panel)

    if store.startswith('マルハン') and not store.endswith('店'):
        store_disp = 'マルハン ' + store[len('マルハン'):] + '店'
    else:
        store_disp = store
    draw.text((W / 2, fy + 30), store_disp,
              font=font(FONT_HEAVY, 32), fill=(255, 90, 140), anchor='mm',
              stroke_width=2, stroke_fill=(255, 255, 255))
    draw.text((W / 2, fy + 62),
              f'♡ 20スロ ／ {footer_date} ／ 差枚ランキング TOP10 ♡',
              font=font(FONT_BOLD, 16), fill=(120, 100, 160), anchor='mm')

    # フッター左右にハート飾り
    draw_heart(draw, 35, fy + 45, 10, (255, 150, 180, 255))
    draw_heart(draw, W - 35, fy + 45, 10, (255, 150, 180, 255))

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    img.convert('RGB').save(out_path, 'PNG', optimize=True)


# ===== ファイル名解析 / 一括処理 =====
# ピーアーク系列店舗一覧（ファイル名は地域名のみのことが多いため、自動で「ピーアーク」を補完）
PIARK_STORES = [
    'ピーアーク北千住',
    'ピーアーク北千住SSS',
    'ピーアーク三田',
    'ピーアーク北綾瀬駅前',
    'ピーアーク竹ノ塚スタジオ',
    'ピーアーク千葉駅前',
    'ピーアーク おゆみ野',
    'ピーアーク 南行徳NEO',
    'ピーアーク朝霞',
    'ピーアーク春日部',
    'ピーアーク新田',
    'ピーアーク松原',
    'ピーアーク越谷',
    'ピーアーク谷塚',
    'ピーアーク草加',
    'ピーアーク相模大野',
    'ピーアーク相模原',
    'ピーアーク新城',
]


def _piark_lookup():
    """{正規化された地域名: 正式店舗名} のマップを返す。"""
    m = {}
    for full in PIARK_STORES:
        loc = full.replace('ピーアーク', '', 1).lstrip(' 　')
        key = unicodedata.normalize('NFKC', loc).replace(' ', '').replace('　', '').lower()
        m[key] = full
    return m


def resolve_store_name(name: str) -> str:
    """店舗名を正規化。
    - すでに「ピーアーク」または既知チェーン名で始まっていればそのまま
    - そうでなければ PIARK_STORES から最長一致で補完
    - 末尾の「店」は除去
    """
    if not name:
        return name
    s = name.strip()
    if s.endswith('店'):
        s = s[:-1]
    # 既にチェーン名が付いていれば触らない
    KNOWN_PREFIXES = ('ピーアーク', 'マルハン', 'メガガイア', 'ガイア', 'ニラク', 'キコーナ')
    if any(s.startswith(p) for p in KNOWN_PREFIXES):
        return s
    # ピーアーク系列の地域名と一致するか
    key = unicodedata.normalize('NFKC', s).replace(' ', '').replace('　', '').lower()
    table = _piark_lookup()
    # 完全一致を優先
    if key in table:
        return table[key]
    # 部分一致（候補が1つに絞れる場合のみ）
    cands = [v for k, v in table.items() if key in k or k in key]
    if len(cands) == 1:
        return cands[0]
    return s


def parse_filename(path):
    base = os.path.splitext(os.path.basename(path))[0]
    # 神の子取材系の店舗名エイリアス（共通）
    STUDIO_ALIAS = {
        'スタジオ': 'ピーアーク竹ノ塚スタジオ',
    }
    # スタジオ形式A: 【店舗】【M.D】【取材名】...
    m = re.match(r'^【([^】]+)】【(\d{1,2})\.(\d{1,2})】', base)
    if m:
        store = STUDIO_ALIAS.get(m.group(1), m.group(1))
        mo, d = int(m.group(2)), int(m.group(3))
        year = datetime.now().year
        return f'{year}{mo:02d}{d:02d}', resolve_store_name(store)
    # スタジオ形式A2: 【店舗】YYYY.M.D... （年付き・区切りドット）
    m = re.match(r'^【([^】]+)】\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})', base)
    if m:
        store = STUDIO_ALIAS.get(m.group(1), m.group(1))
        y, mo, d = int(m.group(2)), int(m.group(3)), int(m.group(4))
        return f'{y}{mo:02d}{d:02d}', resolve_store_name(store)
    # スタジオ形式B: 【店舗】MMDD神の子... または 【店舗】MMDD...
    # ※ A2 より後に評価することで YYYY.M.D を MMDD と誤検出しないようにする
    m = re.match(r'^【([^】]+)】(\d{2})(\d{2})(?![\d.])', base)
    if m:
        store = STUDIO_ALIAS.get(m.group(1), m.group(1))
        mo, d = int(m.group(2)), int(m.group(3))
        year = datetime.now().year
        return f'{year}{mo:02d}{d:02d}', resolve_store_name(store)
    # スタジオ形式C: 【店舗】M.D... （日付に【】なし、区切りはドット、年なし）
    m = re.match(r'^【([^】]+)】\s*(\d{1,2})\.(\d{1,2})(?!\d)', base)
    if m:
        store = STUDIO_ALIAS.get(m.group(1), m.group(1))
        mo, d = int(m.group(2)), int(m.group(3))
        year = datetime.now().year
        return f'{year}{mo:02d}{d:02d}', resolve_store_name(store)
    # スタジオ形式D: 【取材タグ】店舗YYYY.M.D... または 店舗YYYY.M.D...
    # 例: 【サボテン推し】ウエスタン一之江2026.5.19結果
    m = re.match(r'^(?:【[^】]+】)?(.+?)(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})', base)
    if m:
        store = m.group(1).strip(' 　_-')
        y, mo, d = int(m.group(2)), int(m.group(3)), int(m.group(4))
        return f'{y}{mo:02d}{d:02d}', resolve_store_name(store)
    m = re.match(r'^(\d{8})_([^_]+)_', base)
    if m:
        return m.group(1), resolve_store_name(m.group(2))
    # 先頭の日付プレフィックス YYYY[-./:]M[-./:]D を吸収（手書き命名の揺れに対応）
    leading_date = None
    m = re.match(r'^(\d{4})[-./:](\d{1,2})[-./:](\d{1,2})\.?\s*', base)
    if m:
        leading_date = f'{m.group(1)}{int(m.group(2)):02d}{int(m.group(3)):02d}'
        base = base[m.end():]
    m = re.match(r'^(.+?)\(20S\)_(\d{4})-(\d{2})-(\d{2})', base)
    if m:
        return m.group(2) + m.group(3) + m.group(4), resolve_store_name(m.group(1))
    # 店舗名YYYY-MM-DD_全台.csv のような (20S) 無し形式
    m = re.match(r'^(.+?)(\d{4})-(\d{2})-(\d{2})', base)
    if m:
        return m.group(2) + m.group(3) + m.group(4), resolve_store_name(m.group(1).rstrip('_'))
    # 先頭が日付のみで残りが店舗名（例: 2026.4:10.マルハン千葉みなとcsv）
    if leading_date:
        store = re.sub(r'csv$', '', base, flags=re.IGNORECASE).strip(' _.')
        if store:
            return leading_date, resolve_store_name(store)
    raise ValueError(f'unsupported filename pattern: {base}')


def process(path):
    date_full, store = parse_filename(path)
    # 店舗名末尾の "店" を除いて出力先フォルダを統一（xlsx と CSV で表記揺れがあるため）
    if store.endswith('店'):
        store = store[:-1]
    date_md = date_full[4:]
    out_dir = os.path.expanduser(f'~/Desktop/ランキング/神の子/{store}')
    out = os.path.join(out_dir, f'{date_md}_{store}.png')
    ranking = load_ranking(path)
    render_image(ranking, date_full, store, out)
    print(f'OK -> {out}')


def main(argv):
    if len(argv) > 1:
        targets = []
        for a in argv[1:]:
            if any(c in a for c in '*?['):
                targets.extend(glob.glob(a))
            else:
                targets.append(a)
    else:
        print('usage: make_ranking_kamiko.py <xlsx_or_csv> ...', file=sys.stderr)
        sys.exit(1)
    for t in targets:
        try:
            process(t)
        except Exception as e:
            print(f'NG -> {t}: {e}', file=sys.stderr)


if __name__ == '__main__':
    main(sys.argv)
