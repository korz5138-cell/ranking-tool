#!/usr/bin/env python3
"""差枚ランキング画像生成（汎用化版：複数xlsxを一括処理）"""
import os
import re
import sys
import csv
import glob
import openpyxl
from PIL import Image, ImageDraw, ImageFont

def _find_font(*candidates):
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return candidates[-1] if candidates else None


# macOS（ヒラギノ）→ Linux（Noto Sans CJK）の順
FONT_HEAVY = _find_font(
    '/System/Library/Fonts/ヒラギノ角ゴシック W9.ttc',
    '/usr/share/fonts/opentype/noto/NotoSansCJK-Black.ttc',
    '/usr/share/fonts/truetype/noto/NotoSansCJK-Black.ttc',
    '/usr/share/fonts/opentype/noto-cjk/NotoSansCJK-Black.ttc',
)
FONT_BOLD = _find_font(
    '/System/Library/Fonts/ヒラギノ角ゴシック W8.ttc',
    '/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc',
    '/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc',
    '/usr/share/fonts/opentype/noto-cjk/NotoSansCJK-Bold.ttc',
)
FONT_MED = _find_font(
    '/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc',
    '/usr/share/fonts/opentype/noto/NotoSansCJK-Medium.ttc',
    '/usr/share/fonts/truetype/noto/NotoSansCJK-Medium.ttc',
    '/usr/share/fonts/opentype/noto-cjk/NotoSansCJK-Medium.ttc',
)

def font(p, s): return ImageFont.truetype(p, s)


def short_name(n):
    s = str(n)
    for p in ('スマスロ ', 'スマスロ', 'Lパチスロ ', 'L パチスロ ', 'パチスロ ', 'L ', '真打 '):
        if s.startswith(p):
            return s[len(p):]
    return s


def load_ranking_xlsx(xlsx):
    """xlsx 形式: 台番, 機種名(表記), 機種名(正式), 差枚, BB, RB, ART, G数"""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
            if r and r[0] is not None]
    rows.sort(key=lambda r: -(r[3] or 0))
    return [{
        'dai':   int(r[0]),
        'name':  short_name(r[2] if r[2] is not None else r[1]),
        'samai': int(r[3] or 0),
        'bb':    int(r[4] or 0),
        'rb':    int(r[5] or 0),
    } for r in rows[:15]]


def load_ranking_csv(csv_path):
    """CSV 形式: 機種名,台番,総ゲーム,BB,RB,ART,合成確率,差枚数"""
    rows = []
    with open(csv_path, encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for r in reader:
            try:
                dai = int(r.get('台番') or r.get('台番号') or 0)
                sa = int(r.get('差枚数') or r.get('差枚') or 0)
            except ValueError:
                continue
            if dai == 0:
                continue
            def to_int(x):
                try: return int(x)
                except (TypeError, ValueError): return 0
            rows.append({
                'dai':   dai,
                'name':  short_name(r.get('機種名') or ''),
                'samai': sa,
                'bb':    to_int(r.get('BB')),
                'rb':    to_int(r.get('RB')),
            })
    rows.sort(key=lambda x: -x['samai'])
    return rows[:15]


def load_ranking(path):
    if path.lower().endswith('.csv'):
        return load_ranking_csv(path)
    return load_ranking_xlsx(path)


def render_image(ranking, date_full, store, out_path):
    """date_full: 'YYYYMMDD' / store: 店舗名"""
    yyyy = date_full[:4]
    mm = int(date_full[4:6])
    dd = int(date_full[6:8])
    date_badge = f'{mm} / {dd}'
    footer_date = f'{yyyy}年{mm}月{dd}日'

    # ---------- レイアウト ----------
    PAD = 8
    has_bb_rb = bool(ranking) and ('bb' in ranking[0]) and (ranking[0].get('bb') is not None)
    # BB/RB 省略時は機種名カラムを 550 まで広げ、全体幅を維持してタイトルと日付バッジの衝突を防ぐ
    COLS = [
        ('rank', '順位',   100),
        ('dai',  '台番号', 170),
        ('name', '機種名', 320 if has_bb_rb else 550),
        ('sa',   '差枚',   200),
    ]
    if has_bb_rb:
        COLS.append(('bb', 'BB', 115))
        COLS.append(('rb', 'RB', 115))
    W = sum(c[2] for c in COLS) + PAD * 2
    HEADER_H = 200
    ROW_H = 66
    N = len(ranking)
    FOOTER_H = 84
    H = HEADER_H + ROW_H * N + FOOTER_H + 12

    # ---------- 背景 ----------
    img = Image.new('RGB', (W, H), (10, 10, 12))
    draw = ImageDraw.Draw(img)
    for x in range(-H, W + H, 28):
        draw.line([(x, 0), (x - H, H)], fill=(22, 22, 28), width=2)

    # ---------- ヘッダー：タイトル ----------
    ft_title = font(FONT_HEAVY, 64)
    draw.text((80, 88), 'スロット 差枚ランキング', font=ft_title,
              fill=(220, 30, 30), anchor='lm',
              stroke_width=4, stroke_fill=(255, 240, 200))

    # 日付バッジ
    dx, dy, dw, dh = W - 190, 56, 170, 64
    draw.rounded_rectangle([dx, dy, dx + dw, dy + dh], radius=10,
                           fill=(20, 20, 24), outline=(245, 210, 80), width=3)
    draw.text((dx + dw / 2, dy + 16), '集計日', font=font(FONT_BOLD, 16),
              fill=(245, 210, 80), anchor='mm')
    draw.text((dx + dw / 2, dy + 44), date_badge, font=font(FONT_HEAVY, 30),
              fill=(255, 255, 255), anchor='mm')

    # ---------- テーブルヘッダー ----------
    hb_y = HEADER_H - 50
    draw.rectangle([PAD, hb_y, W - PAD, hb_y + 42], fill=(245, 210, 60))
    draw.rectangle([PAD, hb_y + 42, W - PAD, hb_y + 46], fill=(180, 30, 30))
    f_head = font(FONT_HEAVY, 22)
    cx = PAD
    for key, lbl, w in COLS:
        draw.text((cx + w / 2, hb_y + 20), lbl, font=f_head,
                  fill=(20, 20, 20), anchor='mm')
        cx += w
        if cx < W - PAD:
            draw.line([(cx, hb_y + 6), (cx, hb_y + 38)],
                      fill=(120, 80, 0), width=1)

    # ---------- ランクスタイル ----------
    def rank_style(rk):
        if rk == 1: return (220, 35, 35), (255, 220, 80)
        if rk == 2: return (180, 180, 190), (40, 40, 40)
        if rk == 3: return (200, 130, 60), (40, 30, 10)
        return (170, 30, 30), (255, 255, 255)

    # ---------- 行レンダリング ----------
    f_rk     = font(FONT_HEAVY, 36)
    f_rk_lbl = font(FONT_BOLD, 18)
    f_dai    = font(FONT_HEAVY, 34)
    f_dai_lb = font(FONT_BOLD, 14)
    f_nm_l   = [font(FONT_HEAVY, s) for s in (28, 24, 20, 17)]
    f_val    = font(FONT_HEAVY, 32)
    f_unit   = font(FONT_BOLD, 14)
    f_count  = font(FONT_HEAVY, 30)

    def split_two(text, fnt, maxw):
        sep = [i for i, c in enumerate(text) if c in ' 　・-‐ｰー']
        best = None
        for i in sep:
            left = text[:i].rstrip()
            right = text[i + 1:].lstrip()
            if (draw.textlength(left, font=fnt) <= maxw and
                    draw.textlength(right, font=fnt) <= maxw):
                best = (left, right)
        if best:
            return best
        mid = len(text) // 2
        for off in range(0, len(text)):
            for k in (mid - off, mid + off):
                if 0 < k < len(text):
                    L, R = text[:k], text[k:]
                    if (draw.textlength(L, font=fnt) <= maxw and
                            draw.textlength(R, font=fnt) <= maxw):
                        return L, R
        right = text[mid:]
        while draw.textlength(right + '…', font=fnt) > maxw and len(right) > 1:
            right = right[:-1]
        return text[:mid], right + '…'

    y = HEADER_H
    for i, item in enumerate(ranking):
        rk = i + 1
        bg = (22, 22, 28) if i % 2 == 0 else (32, 32, 40)
        draw.rectangle([PAD, y, W - PAD, y + ROW_H - 4], fill=bg)
        if rk <= 3:
            ov = Image.new('RGBA', (W - PAD * 2, ROW_H - 4), (180, 25, 25, 35))
            img.paste(ov, (PAD, y), ov)
        cx = PAD

        # 順位
        w = 100
        bgc, acc = rank_style(rk)
        draw.rounded_rectangle([cx + 8, y + 6, cx + w - 8, y + ROW_H - 10],
                               radius=10, fill=bgc)
        rs = str(rk)
        rsw = draw.textlength(rs, font=f_rk)
        pos_cx = cx + w / 2
        draw.text((pos_cx - 8, y + ROW_H / 2 - 1), rs, font=f_rk,
                  fill=(255, 255, 255), anchor='mm',
                  stroke_width=2, stroke_fill=(80, 0, 0))
        draw.text((pos_cx + rsw / 2 - 4, y + ROW_H / 2 + 6), '位',
                  font=f_rk_lbl, fill=acc, anchor='lm')
        cx += w

        # 台番号
        w = 170
        s = str(item['dai'])
        sw = draw.textlength(s, font=f_dai)
        nx = cx + w / 2 - 14
        draw.text((nx, y + ROW_H / 2 - 1), s, font=f_dai,
                  fill=(255, 255, 255), anchor='mm')
        draw.text((nx + sw / 2 + 4, y + ROW_H / 2 + 8), '番台',
                  font=f_dai_lb, fill=(245, 210, 80), anchor='lm')
        cx += w

        # 機種名
        w = 320 if has_bb_rb else 550
        nm = item['name']
        avail = w - 18
        one_line = None
        for fnt in f_nm_l:
            if draw.textlength(nm, font=fnt) <= avail:
                one_line = fnt
                break
        if one_line is not None:
            draw.text((cx + 12, y + ROW_H / 2), nm, font=one_line,
                      fill=(255, 255, 255), anchor='lm')
        else:
            two_fnt = f_nm_l[-1]
            l1, l2 = split_two(nm, two_fnt, avail)
            draw.text((cx + 12, y + ROW_H / 2 - 14), l1, font=two_fnt,
                      fill=(255, 255, 255), anchor='lm')
            draw.text((cx + 12, y + ROW_H / 2 + 14), l2, font=two_fnt,
                      fill=(255, 255, 255), anchor='lm')
        cx += w

        # 差枚
        w = 200
        sa = item['samai']
        sa_color = (255, 230, 90) if sa >= 0 else (220, 100, 100)
        sa_text = ('+' if sa > 0 else '') + f'{sa:,}'
        mai_x = cx + w - 14
        draw.text((mai_x, y + ROW_H / 2 + 10), '枚', font=f_unit,
                  fill=(245, 210, 80), anchor='rm')
        draw.text((mai_x - 22, y + ROW_H / 2), sa_text, font=f_val,
                  fill=sa_color, anchor='rm')
        cx += w

        # BB
        if not has_bb_rb:
            draw.line([(PAD + 4, y + ROW_H - 4), (W - PAD - 4, y + ROW_H - 4)],
                      fill=(60, 60, 70), width=1)
            y += ROW_H
            continue
        w = 115
        bb_s = str(item['bb'])
        bbw = draw.textlength(bb_s, font=f_count)
        bb_cx = cx + w / 2 - 8
        draw.text((bb_cx, y + ROW_H / 2), bb_s, font=f_count,
                  fill=(255, 255, 255), anchor='mm')
        draw.text((bb_cx + bbw / 2 + 3, y + ROW_H / 2 + 8), '回',
                  font=f_unit, fill=(245, 210, 80), anchor='lm')
        cx += w

        # RB
        w = 115
        rb_s = str(item['rb'])
        rbw = draw.textlength(rb_s, font=f_count)
        rb_cx = cx + w / 2 - 8
        draw.text((rb_cx, y + ROW_H / 2), rb_s, font=f_count,
                  fill=(255, 255, 255), anchor='mm')
        draw.text((rb_cx + rbw / 2 + 3, y + ROW_H / 2 + 8), '回',
                  font=f_unit, fill=(245, 210, 80), anchor='lm')
        cx += w

        draw.line([(PAD + 4, y + ROW_H - 4), (W - PAD - 4, y + ROW_H - 4)],
                  fill=(60, 60, 70), width=1)
        y += ROW_H

    # ---------- フッター ----------
    fy = y + 6
    draw.rectangle([0, fy, W, fy + FOOTER_H], fill=(160, 25, 25))
    draw.rectangle([0, fy, W, fy + 4], fill=(245, 210, 60))
    # 店舗名見出し（「マルハン千葉みなと」→「マルハン 千葉みなと店」）
    if store.startswith('マルハン') and not store.endswith('店'):
        store_disp = 'マルハン ' + store[len('マルハン'):] + '店'
    else:
        store_disp = store
    draw.text((W / 2, fy + 30), store_disp,
              font=font(FONT_HEAVY, 32), fill=(255, 255, 255), anchor='mm')
    draw.text((W / 2, fy + 60),
              f'20スロ ／ {footer_date} ／ 差枚ランキング TOP{N}',
              font=font(FONT_BOLD, 17), fill=(255, 230, 180), anchor='mm')

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    img.save(out_path, 'PNG', optimize=True)


def parse_filename(path):
    """ファイル名から (date_full='YYYYMMDD', store) を抽出。
       対応形式:
         (A) YYYYMMDD_店舗名_20S.xlsx           → date='YYYYMMDD', store=店舗名
         (B) 店舗名(20S)_YYYY-MM-DD_全台.csv    → 店舗名のみ抽出後、日付はハイフン除去
    """
    base = os.path.splitext(os.path.basename(path))[0]
    # (A)
    m = re.match(r'^(\d{8})_([^_]+)_', base)
    if m:
        return m.group(1), m.group(2)
    # (B)
    m = re.match(r'^(.+?)\(20S\)_(\d{4})-(\d{2})-(\d{2})', base)
    if m:
        store = m.group(1)
        date_full = m.group(2) + m.group(3) + m.group(4)
        return date_full, store
    raise ValueError(f'unsupported filename pattern: {base}')


def process(path):
    date_full, store = parse_filename(path)
    date_md = date_full[4:]

    out_dir = os.path.expanduser(f'~/Desktop/ランキング/{store}')
    out     = os.path.join(out_dir, f'{date_md}_{store}.png')

    ranking = load_ranking(path)
    render_image(ranking, date_full, store, out)
    print(f'OK -> {out}')


def main(argv):
    if len(argv) > 1:
        targets = []
        for a in argv[1:]:
            if any(c in a for c in '*?['):
                targets.extend(glob.glob(a))
            else:
                targets.append(a)
    else:
        # 引数なしならデフォルトで千葉みなとフォルダを一括
        targets = sorted(glob.glob(
            os.path.expanduser('~/Desktop/千葉みなと/*_*_20S.xlsx')))

    for t in targets:
        try:
            process(t)
        except Exception as e:
            print(f'NG -> {t}: {e}', file=sys.stderr)


if __name__ == '__main__':
    main(sys.argv)
